# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, reverse
from .models import Pais, Ciudad, Calle, Voto 
from django.db.models import Count
from operator import attrgetter
from django.db.models import Q
import user_agents
import matplotlib.pyplot as plt
from django.utils.html import conditional_escape, format_html_join
import re
from django.core.validators import validate_email
from django.shortcuts import render
from .forms import CommentForm
from .models import Comment, Visit
from django.utils import timezone
from django.urls import reverse
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.contrib.sites.shortcuts import get_current_site
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.models import User
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.conf import settings
import os
from sendgrid.helpers.mail import Mail
import uuid
import smtplib
from email.mime.text import MIMEText


def generate_confirmation_token():
    return str(uuid.uuid4())

def confirmar_voto(request, confirmation_token):
    try:
        voto = Voto.objects.get(confirmation_token=confirmation_token, confirmado=False)
        voto.confirmado = True
        voto.save()
        mensaje = "¡Tu voto ha sido confirmado correctamente!"
        return redirect('ingresar_datos_votante', token=voto.confirmation_token)
    except Voto.DoesNotExist:
        mensaje = "El enlace de confirmación no es válido."
    
    return render(request, 'confirmar_voto.html', {'mensaje': mensaje})

def ingresar_datos_votante(request, token):
    try:
        voto = Voto.objects.get(confirmation_token=token, confirmado=False)
    except Voto.DoesNotExist:
        error_msg = 'Token de confirmación inválido.'
        return render(request, 'error.html', {'error': error_msg})

    if request.method == 'POST':
        # Obtener los datos ingresados por el votante
        email = request.POST.get('email')
        dni = request.POST.get('dni')
        telefono = request.POST.get('telefono')
        nombre_persona = request.POST.get('nombre_persona')
        apellido = request.POST.get('apellido')

        # Actualizar el voto confirmado con los nuevos datos ingresados
        voto.email = email
        voto.dni = dni
        voto.telefono = telefono
        voto.nombre_persona = nombre_persona
        voto.apellido = apellido
        voto.confirmado = True
        voto.save()

        success_msg = 'Tu voto ha sido confirmado exitosamente.'
        return render(request, 'ingresar_datos_votante.html', {'mensaje': success_msg, 'token': token})

    else:
        return render(request, 'ingresar_datos_votante.html', {'voto': voto})


import ssl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr


def enviar_email_confirmacion_voto(voto):
    print('Enviando correo electrónico')
    # Configuración del servidor de correo electrónico
    smtp_host = 'smtp.hostinger.com'
    smtp_ports = [465, 587, 25]
    smtp_username = 'admin@calles.com'
    smtp_password = 'Megusta122022$'
    EMAIL_USE_TLS = True

    # Construir el mensaje de correo electrónico
    mensaje = MIMEMultipart()
    mensaje['From'] = formataddr(('Remitente', '< CiudadanosPorCalles >'' admin@calles.com'))
    mensaje['To'] = voto.email
    mensaje['Subject'] = 'Confirmación de voto'

    # Contenido del mensaje
    contenido = f'''
    <html>
    <body>
        <div style="text-align: center;">
            <h1 style="color: #4CAF50;">¡Gracias por tu voto!</h1>
            <p>Hola {voto.nombre_persona},</p>
            <p>Gracias por tu voto. Para confirmar tu voto, haz clic en el siguiente enlace:</p>
            <p><a href="http://192.168.1.36:8000/ingresar-datos-votante/{voto.token}" style="padding: 10px 20px; background-color: #4CAF50; color: white; text-decoration: none; border-radius: 5px;">Confirmar Voto</a></p>
            <p>¡Gracias y que tengas un buen día!</p>
        </div>
    </body>
    </html>
    '''

    mensaje.attach(MIMEText(contenido, 'html'))

    for smtp_port in smtp_ports:
        try:
            # Intentar iniciar conexión con el servidor SMTP y enviar el correo electrónico
            with smtplib.SMTP(smtp_host, smtp_port) as servidor_smtp:
                if smtp_port in [587, 25]:
                    servidor_smtp.starttls()
                servidor_smtp.login(smtp_username, smtp_password)
                servidor_smtp.send_message(mensaje)
            print('Correo electrónico enviado exitosamente')
            break  # Salir del bucle si el envío fue exitoso
        except smtplib.SMTPException as e:
            # Manejar la excepción y continuar con el siguiente puerto SMTP
            print(f'Error al enviar el correo electrónico en el puerto {smtp_port}: {str(e)}')
            continue

    else:
        # El bucle for se completó sin éxito en el envío del correo electrónico
        print('No se pudo enviar el correo electrónico en ninguno de los puertos SMTP disponibles')
        # Realizar acciones alternativas, como registrar el error en un archivo de registro o notificar al administrador



def votar(request, calle_id):
    # Obtener la calle, ciudad y país
    calle = Calle.objects.get(id=calle_id)
    ciudad = calle.ciudad
    pais = ciudad.pais  

    # Verificar si la votación ya ha comenzado
    if timezone.now() < calle.fecha_inicio:
        error_msg = 'La votación aún no ha comenzado.'
        return render(request, 'votar.html', {'error': error_msg})


    # Verificar si la fecha actual es menor o igual a la fecha de finalización de la calle
    if timezone.now() > calle.fecha_finalizacion:
        error_msg = 'La votación para esta calle ha finalizado.'
        return render(request, 'votar.html', {'calle': calle, 'error': error_msg})


    if request.method == 'POST':
        email = request.POST.get('email')
        dni = request.POST.get('dni')
        telefono = request.POST.get('telefono')
        nombre_persona = request.POST.get('nombre_persona')
        apellido = request.POST.get('apellido')

        # Validar el email con regex
        if not re.match(r'^[\w\.-]+@[\w\.-]+\.[\w]{2,}$', email):
            error_msg = 'El email ingresado no es válido.'
            return render(request, 'votar.html', {'calle': calle, 'ciudad': ciudad, 'pais': pais, 'error': error_msg})



        # Validar el DNI (solo números, sin puntos ni guiones)
        if not re.match(r'^\d{8}$', dni):
            error_msg = 'El DNI debe contener solo 8 dígitos.'
            return render(request, 'votar.html', {'calle': calle, 'ciudad': ciudad, 'pais': pais, 'error': error_msg})

        # Validar el número de teléfono (solo números, sin paréntesis ni guiones)
        if not re.match(r'^\d{10}$', telefono):
            error_msg = 'El número de teléfono debe contener solo 10 dígitos.'
            return render(request, 'votar.html', {'calle': calle, 'ciudad': ciudad, 'pais': pais, 'error': error_msg})

        # Verificar si el votante ya ha votado con la misma información
        if Voto.objects.filter(Q(email=email) | Q(dni=dni) | Q(telefono=telefono)).exists():
            error_msg = 'Ya has votado con esta información.'
            return render(request, 'votar.html', {'calle': calle, 'ciudad': ciudad, 'pais': pais, 'error': error_msg})



        # Si el votante no ha votado aún, crear y guardar el voto
        #voto = Voto.objects.create(email=email, dni=dni, telefono=telefono, calle=calle, pais=pais, ciudad=ciudad, nombre_persona=nombre_persona, apellido=apellido)
        #voto.save()


        calle.cantidad_votos = Voto.objects.filter(calle=calle).count()
        calle.save()

        votos_por_calle = calle.cantidad_votos


        # Crear un archivo Excel y escribir los datos
        headers = ['Email', 'DNI', 'Teléfono', 'Cantidad de votos de la calle', 'Ciudad', 'País', 'Nombre', 'Apellido', 'Voto']
        filename = 'votos.xlsx'

        try:
            # Leer el archivo Excel existente
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            # Si el archivo Excel no existe, crear uno nuevo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(headers)


           # Obtener la cantidad de votos de la calle
            votos_por_calle = Voto.objects.filter(calle=calle, confirmado=True).count()


            # Agregar los nuevos datos a las celdas correspondientes
            row = [email, dni, telefono, votos_por_calle, ciudad.nombre, pais.nombre, nombre_persona, apellido, "Voto"]
            ws.append(row)

            # Guardar el archivo Excel
            wb.save(filename)


        # Crear un nuevo voto
        token = str(uuid.uuid4())
        voto = Voto.objects.create(
            calle=calle,
            ciudad=ciudad,
            pais=pais,
            email=email,
            dni=dni,
            telefono=telefono,
            nombre_persona=nombre_persona,
            apellido=apellido,
            token=token,
            confirmation_token=token   
        )
       
        print(f"Token generado: {voto.token}")
        print(f"Token almacenado en el campo confirmation_token: {voto.confirmation_token}")
          
        # Enviar el email de confirmación con el enlace correspondiente
        enviar_email_confirmacion_voto(voto)

        request.session['correo_enviado'] = True

        success_msg = 'Se ha enviado un email de confirmación a tu dirección de correo electrónico.'
        return render(request, 'gracias.html', {'calle': calle, 'ciudad': ciudad, 'pais': pais, 'success': success_msg})

    else:
        return render(request, 'votar.html', {'calle': calle, 'ciudad': ciudad, 'pais': pais})





def procesar_voto(request, token):
    try:
        voto = Voto.objects.get(token=token)
    except Voto.DoesNotExist:
        error_msg = 'Token de votación inválido.'
        return render(request, 'error.html', {'error': error_msg})
    
    if request.method == 'POST':
        # Actualizar el campo 'confirmado' a True
        voto.confirmado = True
        voto.save()

        mensaje_confirmacion = '¡Tu voto ha sido confirmado correctamente!'

        return render(request, 'confirmar_voto.html', {'mensaje': mensaje_confirmacion})
    else:
        return render(request, 'votar.html')
    


    

def error_page(request):
    error_msg = 'Error desconocido.'  # Mensaje de error predeterminado

    # Verificar si se ha pasado un mensaje de error personalizado en la URL
    if 'error_msg' in request.GET:
        error_msg = request.GET['error_msg']

    return render(request, 'error.html', {'error': error_msg})




def paises(request):
    return render(request, 'paises.html')
    


def ciudades(request, pais_id):
    pais = Pais.objects.get(id=pais_id)
    ciudades = Ciudad.objects.filter(pais=pais)
    return render(request, 'ciudades.html', {'pais': pais, 'ciudades': ciudades})



def calles(request, ciudad_id):
    ciudad = Ciudad.objects.get(pk=ciudad_id)
    calles = Calle.objects.filter(ciudad=ciudad).annotate(num_votos=Count('voto'))
    calles_ordenadas = sorted(calles, key=attrgetter('num_votos'), reverse=True)
    return render(request, 'calles.html', {'ciudad': ciudad, 'calles': calles_ordenadas})


def listado_votos(request):
    votos = Voto.objects.all()
    votos_count = {}
    for voto in votos:
        pais = voto.pais.nombre
        ciudad = voto.ciudad.nombre
        calle = voto.calle.nombre
        key = (pais, ciudad, calle)
        if key in votos_count:
            votos_count[key] += 1
        else:
            votos_count[key] = 1
    votos_list = []
    for key, value in votos_count.items():
        pais, ciudad, calle = key
        votos_list.append((pais, ciudad, calle, value))
    
    votos_list = sorted(votos_list, key=lambda x: x[3], reverse=True)  # Ordenar de mayor a menor
    
    return render(request, 'listado_votos.html', {'votos_list': votos_list})



#def listado_votos(request):
#    votos = Voto.objects.all()
#    return render(request, 'listado_votos.html', {'votos': votos})


def lista_paises(request):
    paises = Pais.objects.all()
    return render(request, 'paises.html', {'paises': paises})


def gracias(request):
    return render(request, 'gracias.html')


def error_404(request, exception):
    return render(request, '404.html', status=404)

def error_500(request):
    return render(request, '500.html', status=500)

def detalle_calle(request, calle_id):
    calle = Calle.objects.get(pk=calle_id)
    return render(request, 'detalles_calle.html', {'calle': calle})



def home(request):
    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            message = form.cleaned_data['message']

            # Validar el nombre (solo letras de la a a la z sin caracteres especiales ni números)
            if not re.match(r'^[a-zA-Z ]+$', name):
                error_msg = 'El nombre solo debe contener letras.'
                return render(request, 'home.html', {'form': form, 'error': error_msg})

            # Validar el email con regex
            if not re.match(r'^[\w\.-]+@[\w\.-]+\.[\w]{2,}$', email):
                error_msg = 'El email ingresado no es válido.'
                return render(request, 'home.html', {'form': form, 'error': error_msg})

            # Validar el mensaje (solo letras, números y signos de puntuación básicos)
            if not re.match(r'^[a-zA-Z0-9 .,;:!?"\'-]+$', message):
                error_msg = 'El mensaje contiene caracteres no permitidos.'
                return render(request, 'home.html', {'form': form, 'error': error_msg})


            # Guardar el comentario en la base de datos
            comment = Comment(name=name, email=email, message=message)
            comment.save()
            return redirect('home')
    else:
        form = CommentForm()
    # Obtener todos los comentarios almacenados en la base de datos
    comments = Comment.objects.all()

    # Votos por calle
    #votos_por_calle = Voto.objects.values('calle__nombre').annotate(num_votos=Count('id'))
    # Votos por calle
    votos_por_calle = Voto.objects.filter(confirmado=True).values('calle__nombre').annotate(num_votos=Count('id'))
    voto_por_ciudad = Voto.objects.filter(confirmado=True).values('ciudad__nombre').annotate(num_votos=Count('id'))
    voto_por_pais = Voto.objects.filter(confirmado=True).values('pais__nombre').annotate(num_votos=Count('id'))

    # Obtiene la dirección IP real del visitante
    ip_address = request.META.get('HTTP_X_REAL_IP', '')
    if not ip_address:
        ip_address = request.META.get('HTTP_X_FORWARDED_FOR', '')
        if ip_address:
            # La cabecera 'X-Forwarded-For' contiene una lista de direcciones IP,
            # la dirección IP real del visitante suele ser la última dirección en la lista.
            ip_address = ip_address.split(',')[-1].strip()

    user_agent = request.META.get('HTTP_USER_AGENT', '')
    visit = Visit(ip_address=ip_address)

    if user_agent:
        ua = user_agents.parse(user_agent)
        visit.os = ua.os.family
        visit.browser = ua.browser.family
        visit.device = ua.device.family

    visit.save()

    context = {'form': form, 'comments': comments ,'votos_por_calle': votos_por_calle,'voto_por_ciudad': voto_por_ciudad ,'voto_por_pais': voto_por_pais}
    return render(request, 'home.html', context)
